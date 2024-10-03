import base64
import os
from dotenv import load_dotenv
import pandas as pd
import openpyxl
from pathlib import Path
from openpyxl import Workbook
from xlrd import open_workbook
from io import BytesIO
from openpyxl_image_loader import SheetImageLoader
 
load_dotenv()

def convert_xls_to_xlsx(xls_path):
    # Open the .xls file
    with open_workbook(xls_path) as xls_book:
        # Create a new .xlsx workbook
        xlsx_book = Workbook()

        # Iterate through all sheets
        for i in range(xls_book.nsheets):
            # Get the worksheet
            xls_sheet = xls_book.sheet_by_index(i)
            
            # Create a new worksheet in xlsx file with the same name
            if i == 0:
                xlsx_sheet = xlsx_book.active
                xlsx_sheet.title = xls_sheet.name
            else:
                xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)

            # Copy the cell values and styles
            for row in range(xls_sheet.nrows):
                for col in range(xls_sheet.ncols):
                    xlsx_sheet.cell(row=row+1, column=col+1, value=xls_sheet.cell_value(row, col))

        # Generate the .xlsx filename
        xlsx_path = os.path.splitext(xls_path)[0] + '.xlsx'

        # Save the .xlsx file
        xlsx_book.save(xlsx_path)

    return xlsx_path

def images_xlsx(sheet, header_row, img_col):
    loader = SheetImageLoader(sheet)
    images = []
    for row in sheet.iter_rows(min_row=header_row + 2, max_row=sheet.max_row, min_col=img_col + 1, max_col=img_col + 1):
        for cell in row:
            try:
                image = loader.get(cell.coordinate)
                img_byte_arr = BytesIO()
                image.save(img_byte_arr, format='png')
                images.append(base64.b64encode(img_byte_arr.getvalue()))
            except:
                images.append(None)
    return images

def df_maker(path):
    # Checking if a .xlsx file already exists
    file_check = Path(''.join(path.split('.')[:-1]) + ".xlsx")
    if not file_check.is_file():
        print("Converting file to .xlsx format...")
        path = convert_xls_to_xlsx(path)
    else:
        path = ''.join(path.split('.')[:-1]) + ".xlsx"
    print(path)

    excel_obj = pd.ExcelFile(path)
    sheets = []
    names = []
    for sheet in excel_obj.book:
        # Getting the index of header row
        for idx, row in enumerate(sheet):
            header_cols = ['qty', 'price', 'quantity', 'description']
            if any([val in ''.join([str(row_str.value) for row_str in row]).lower() for val in header_cols]):
                header_row = idx
                break
        # Getting the sheet name and loading the sheet
        try:
            sheet_name = str(sheet.title)
        except:
            try:
                sheet_name = str(sheet).split(':')[1].replace('<', '').replace('>', '')
            except Exception as e:
                print(e)
        print(sheet_name)
        if sheet_name != "Summary" and sheet_name != "Export Summary":
            df = pd.read_excel(excel_obj, sheet_name, header=header_row)
            names.append(sheet_name)
            # Getting the image column
            img_cols, col_name = [], []
            for idx, key in enumerate(df.keys()):
                if 'image' in str(key).lower():
                    img_cols.append(idx)
                    col_name.append(key)

            xlsx_file = openpyxl.load_workbook(path)
            img_sheet = xlsx_file[sheet_name]

            # Adding the images to their respective columns
            if col_name is not None:
                for img_col, col in zip(img_cols, col_name):
                    images = images_xlsx(img_sheet, header_row, img_col)
                    df[col] = images #list(map(lambda x: x.decode(), images))
            sheets.append(df)
    return sheets, names
    
# if __name__ == "__main__":
#     path = r"S:\AutoQuote\data\WALTHR PRICE LIST.xls"
#     df = df_maker(path)
#     df[0].to_excel('multiple_images.xlsx')
#     print("File saved!")